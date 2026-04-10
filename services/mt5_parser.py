"""Parser for MT5 Excel report files (.xlsx).

MT5 reports contain multiple tables with header rows. We locate the
"Positions" table (or "Positions" / "Deals" depending on language),
extract trades and map them to our internal format.

Expected columns in the Positions table:
  Heure | Position | Symbole | Type | Volume | Prix | S/L | T/P |
  Heure (close) | Prix (close) | Commission | Echange | Profit
"""

import io
import re
from datetime import datetime

import pandas as pd


# Possible section headers that mark the positions table
_POSITIONS_MARKERS = {"positions", "deals", "orders"}

# MT5 column name variants (lowercased) → internal field
# The MT5 format has duplicate column names (Heure, Prix appear twice:
# once for open, once for close). We handle this by column index.
_MT5_TYPE_MAP = {
    "buy": "long",
    "sell": "short",
}


def _find_positions_table(wb_data: bytes) -> pd.DataFrame | None:
    """Scan the Excel file for the Positions table and return it as a DataFrame."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(wb_data), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None

    # Step 1: Find the row that contains the "Positions" section marker
    positions_start = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and str(cell).strip().lower() in _POSITIONS_MARKERS:
                positions_start = i
                break
        if positions_start is not None:
            break

    if positions_start is None:
        return None

    # Step 2: Find the header row (first row after marker with content in
    # multiple columns). The header row is typically right after the marker
    # or the marker row itself contains headers.
    header_idx = None
    for i in range(positions_start, min(positions_start + 5, len(rows))):
        row = rows[i]
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 5:
            # This looks like a header row — check it contains expected fields
            lower_vals = {str(c).strip().lower() for c in row if c}
            if lower_vals & {"type", "volume", "symbole", "symbol", "profit"}:
                header_idx = i
                break

    if header_idx is None:
        return None

    headers = list(rows[header_idx])

    # Step 3: Collect data rows until we hit an empty row or a new section
    data_rows = []
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) < 3:
            break
        data_rows.append(list(row))

    if not data_rows:
        return None

    df = pd.DataFrame(data_rows, columns=range(len(headers)))
    # Attach raw headers for mapping
    df.attrs["raw_headers"] = headers
    return df


def _resolve_columns(headers: list) -> dict[str, int]:
    """Map internal field names to column indices based on MT5 header names.

    MT5 has duplicate column names (Heure/Prix appear for both open and close).
    We use positional logic: the first occurrence is open, the second is close.
    """
    mapping = {}
    lower_headers = [str(h).strip().lower() if h else "" for h in headers]

    # Simple 1:1 mappings
    simple = {
        "position": "ticket",
        "symbole": "symbol", "symbol": "symbol",
        "type": "type",
        "volume": "volume",
        "commission": "commission",
        "echange": "swap", "swap": "swap",
        "profit": "profit", "bénéfice": "profit",
    }

    for i, h in enumerate(lower_headers):
        if h in simple:
            field = simple[h]
            if field not in mapping:
                mapping[field] = i

    # Duplicate columns: Heure (open_time, close_time), Prix (entry_price, exit_price)
    time_indices = [i for i, h in enumerate(lower_headers) if h in ("heure", "time", "date", "open time", "close time")]
    price_indices = [i for i, h in enumerate(lower_headers) if h in ("prix", "price", "open price", "close price")]

    if len(time_indices) >= 2:
        mapping["open_time"] = time_indices[0]
        mapping["close_time"] = time_indices[1]
    elif len(time_indices) == 1:
        mapping["open_time"] = time_indices[0]

    if len(price_indices) >= 2:
        mapping["entry_price"] = price_indices[0]
        mapping["exit_price"] = price_indices[1]
    elif len(price_indices) == 1:
        mapping["entry_price"] = price_indices[0]

    # S/L and T/P (informational, not needed for trades — skip)

    return mapping


def _parse_mt5_datetime(val) -> datetime:
    """Parse MT5 datetime which can be string or datetime object."""
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    # MT5 format: 2025.12.01 13:20:45
    for fmt in ("%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Fallback to pandas
    return pd.to_datetime(s).to_pydatetime()


def _parse_mt5_number(val) -> float:
    """Parse a number that may use space as thousands separator and comma as decimal."""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    # Remove spaces (thousands separator)
    s = s.replace("\u00a0", "").replace(" ", "")
    # Handle comma as decimal separator (European format)
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "," in s and "." in s:
        # e.g. 4.249,73 → remove dots, replace comma
        s = s.replace(".", "").replace(",", ".")
    return float(s)


def parse_mt5_excel(
    file_content: bytes,
) -> tuple[list[dict], list[str], float | None, str | None]:
    """Parse an MT5 Excel report and return trades.

    Returns:
        Tuple (trades, errors, detected_balance, detected_currency).
    """
    # Try to detect currency from the account info header
    detected_currency = _detect_currency_from_header(file_content)

    df = _find_positions_table(file_content)
    if df is None:
        return [], ["Tableau 'Positions' introuvable dans le fichier Excel"], None, detected_currency

    headers = df.attrs.get("raw_headers", [])
    col_map = _resolve_columns(headers)

    # Validate required columns
    required = {"open_time", "close_time", "symbol", "type", "volume", "entry_price", "exit_price", "profit"}
    missing = required - set(col_map.keys())
    if missing:
        return [], [f"Colonnes manquantes : {', '.join(missing)}"], None, detected_currency

    trades: list[dict] = []
    errors: list[str] = []

    for idx in range(len(df)):
        row = df.iloc[idx]
        line_num = idx + 1
        try:
            type_val = str(row[col_map["type"]]).strip().lower()

            # Skip non-trade rows (balance, credit, etc.)
            if type_val not in _MT5_TYPE_MAP:
                continue

            side = _MT5_TYPE_MAP[type_val]
            open_time = _parse_mt5_datetime(row[col_map["open_time"]])
            close_time = _parse_mt5_datetime(row[col_map["close_time"]])
            symbol = str(row[col_map["symbol"]]).strip()
            entry_price = _parse_mt5_number(row[col_map["entry_price"]])
            exit_price = _parse_mt5_number(row[col_map["exit_price"]])
            volume = _parse_mt5_number(row[col_map["volume"]])
            profit = _parse_mt5_number(row[col_map["profit"]])

            # Include commission + swap in PnL if available
            commission = 0.0
            swap = 0.0
            if "commission" in col_map and row[col_map["commission"]] is not None:
                try:
                    commission = _parse_mt5_number(row[col_map["commission"]])
                except (ValueError, TypeError):
                    pass
            if "swap" in col_map and row[col_map["swap"]] is not None:
                try:
                    swap = _parse_mt5_number(row[col_map["swap"]])
                except (ValueError, TypeError):
                    pass

            total_pnl = profit + commission + swap

            trades.append({
                "open_time": open_time,
                "close_time": close_time,
                "symbol": symbol,
                "side": side,
                "entry_price": entry_price,
                "exit_price": exit_price,
                "lot_size": volume,
                "pnl": total_pnl,
                "pips": None,
            })

        except Exception as e:
            errors.append(f"Ligne {line_num}: {e}")

    detected_balance = _detect_initial_deposit(file_content)

    return trades, errors, detected_balance, detected_currency


def preview_mt5_excel(
    file_content: bytes,
    max_trades: int = 10,
) -> dict:
    """Preview MT5 Excel parsing: column mapping status + first N trades.

    Returns dict with keys: columns, trades, total, errors, currency, raw_headers.
    """
    detected_currency = _detect_currency_from_header(file_content)

    df = _find_positions_table(file_content)
    if df is None:
        return {
            "error": "Tableau 'Positions' introuvable dans le fichier Excel",
            "columns": [],
            "trades": [],
            "total": 0,
            "errors": ["Tableau 'Positions' introuvable"],
            "currency": detected_currency,
            "raw_headers": [],
        }

    headers = df.attrs.get("raw_headers", [])
    col_map = _resolve_columns(headers)
    raw_headers = [str(h).strip() for h in headers if h]

    # Build column mapping status
    _expected = [
        ("open_time", "Open Time", True),
        ("close_time", "Close Time", True),
        ("symbol", "Symbol", True),
        ("type", "Side (Type)", True),
        ("volume", "Lot Size", True),
        ("entry_price", "Entry Price", True),
        ("exit_price", "Exit Price", True),
        ("profit", "Profit", True),
        ("commission", "Commission", False),
        ("swap", "Swap", False),
        ("ticket", "Ticket / Position", False),
    ]

    columns_status = []
    for field, label, required in _expected:
        found = field in col_map
        source = str(headers[col_map[field]]).strip() if found else None
        columns_status.append({
            "field": field,
            "label": label,
            "found": found,
            "required": required,
            "source_column": source,
        })

    # Check required columns
    required_fields = {"open_time", "close_time", "symbol", "type", "volume", "entry_price", "exit_price", "profit"}
    missing = required_fields - set(col_map.keys())
    if missing:
        return {
            "columns": columns_status,
            "trades": [],
            "total": 0,
            "errors": [f"Colonnes manquantes : {', '.join(missing)}"],
            "currency": detected_currency,
            "raw_headers": raw_headers,
        }

    trades: list[dict] = []
    errors: list[str] = []
    total_count = 0

    for idx in range(len(df)):
        row = df.iloc[idx]
        line_num = idx + 1
        try:
            type_val = str(row[col_map["type"]]).strip().lower()
            if type_val not in _MT5_TYPE_MAP:
                continue
            total_count += 1

            if len(trades) >= max_trades:
                continue  # keep counting total

            side = _MT5_TYPE_MAP[type_val]
            open_time = _parse_mt5_datetime(row[col_map["open_time"]])
            close_time = _parse_mt5_datetime(row[col_map["close_time"]])
            symbol = str(row[col_map["symbol"]]).strip()
            entry_price = _parse_mt5_number(row[col_map["entry_price"]])
            exit_price = _parse_mt5_number(row[col_map["exit_price"]])
            volume = _parse_mt5_number(row[col_map["volume"]])
            profit = _parse_mt5_number(row[col_map["profit"]])

            commission = 0.0
            swap = 0.0
            if "commission" in col_map and row[col_map["commission"]] is not None:
                try:
                    commission = _parse_mt5_number(row[col_map["commission"]])
                except (ValueError, TypeError):
                    pass
            if "swap" in col_map and row[col_map["swap"]] is not None:
                try:
                    swap = _parse_mt5_number(row[col_map["swap"]])
                except (ValueError, TypeError):
                    pass

            total_pnl = profit + commission + swap

            trades.append({
                "open_time": open_time.strftime("%Y-%m-%d %H:%M:%S"),
                "close_time": close_time.strftime("%Y-%m-%d %H:%M:%S"),
                "symbol": symbol,
                "side": side,
                "entry_price": entry_price,
                "exit_price": exit_price,
                "lot_size": volume,
                "pnl": round(total_pnl, 2),
            })
        except Exception as e:
            errors.append(f"Ligne {line_num}: {e}")

    detected_balance = _detect_initial_deposit(file_content)

    return {
        "columns": columns_status,
        "trades": trades,
        "total": total_count,
        "errors": errors,
        "currency": detected_currency,
        "initial_balance": detected_balance,
        "raw_headers": raw_headers,
    }


def _detect_initial_deposit(file_content: bytes) -> float | None:
    """Scan the Transactions section for the 'Initial Deposit' row.

    MT5 Report History structure:
      Row N  : 'Transactions' / 'Deals' / 'Opérations' (section marker)
      Row N+1: headers (Heure | Opération | ... | Profit | Solde | Commentaire)
      Row N+2+: data — look for row containing 'initial' and 'deposit'/'dépôt'
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the Transactions section
    _TX_MARKERS = {"transactions", "deals", "opérations", "operations"}
    tx_start = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and str(cell).strip().lower() in _TX_MARKERS:
                tx_start = i
                break
        if tx_start is not None:
            break

    if tx_start is None:
        return None

    # Find the header row and locate Solde/Balance + Profit columns
    header_row = rows[tx_start + 1] if tx_start + 1 < len(rows) else None
    if header_row is None:
        return None

    lower_headers = [str(h).strip().lower() if h else "" for h in header_row]
    solde_idx = None
    profit_idx = None
    for j, h in enumerate(lower_headers):
        if h in ("solde", "balance"):
            solde_idx = j
        if h in ("profit", "bénéfice"):
            profit_idx = j

    target_idx = solde_idx if solde_idx is not None else profit_idx
    if target_idx is None:
        return None

    # Scan data rows after the header for Initial Deposit / Dépôt initial
    for i in range(tx_start + 2, min(tx_start + 20, len(rows))):
        row = rows[i]
        row_strs = [str(c).strip().lower().replace("+", " ") for c in row if c is not None]
        comment = " ".join(row_strs)
        is_deposit = (
            ("initial" in comment and "deposit" in comment)
            or ("dépôt" in comment and "initial" in comment)
            or ("depot" in comment and "initial" in comment)
            or "initial deposit" in comment
            or "dépôt initial" in comment
        )
        if is_deposit:
            val = row[target_idx]
            if val is not None:
                try:
                    return float(val) if isinstance(val, (int, float)) else _parse_mt5_number(str(val))
                except (ValueError, TypeError):
                    pass
    return None


def _detect_currency_from_header(file_content: bytes) -> str | None:
    """Scan the first rows of the Excel for account info containing currency."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    # Check first 10 rows for currency info
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            # Pattern: "Compte: 13857368 (USD, ...)" or "Account: ... (USD, ...)"
            m = re.search(r'\((\w{3}),', s)
            if m:
                currency = m.group(1).upper()
                if currency in ("USD", "EUR", "GBP", "JPY", "CHF", "AUD", "CAD", "NZD"):
                    wb.close()
                    return currency
    wb.close()
    return None
