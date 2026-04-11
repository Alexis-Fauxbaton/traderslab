import json
import math

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session

from database import get_db
from models.run import Run
from models.trade import Trade
from models.variant import Variant
from models.strategy import Strategy
from models.user import User
from schemas.run import RunOut, RunDetail, RunImportResponse, TradesPaginated
from schemas.trade import TradeOut
from services.csv_parser import parse_csv
from services.mt5_parser import parse_mt5_excel, preview_mt5_excel
from services.metrics import compute_metrics, _compute_sharpe_annualized
from services.aggregation import recompute_variant_metrics, recompute_strategy_metrics, recompute_run_metrics, _run_metrics_stale
from services.auth import get_current_user

router = APIRouter(prefix="/runs", tags=["runs"])


def _verify_run_owner(run_id: str, db: Session, user: User) -> Run:
    """Get run and verify ownership through variant → strategy."""
    run = db.query(Run).filter(Run.id == run_id).first()
    if not run:
        raise HTTPException(404, "Run introuvable")
    variant = db.query(Variant).filter(Variant.id == run.variant_id).first()
    if not variant:
        raise HTTPException(404, "Run introuvable")
    strategy = db.query(Strategy).filter(Strategy.id == variant.strategy_id, Strategy.user_id == user.id).first()
    if not strategy:
        raise HTTPException(404, "Run introuvable")
    return run


@router.get("", response_model=list[RunOut])
def list_runs(variant_id: str = Query(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Verify variant ownership
    variant = db.query(Variant).filter(Variant.id == variant_id).first()
    if variant:
        strategy = db.query(Strategy).filter(Strategy.id == variant.strategy_id, Strategy.user_id == current_user.id).first()
        if not strategy:
            raise HTTPException(404, "Variante introuvable")
    return (
        db.query(Run)
        .filter(Run.variant_id == variant_id)
        .order_by(Run.imported_at.desc())
        .all()
    )


@router.get("/{run_id}/summary", response_model=RunOut)
def get_run_summary(run_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Retourne le run sans ses trades — léger, pour header + métriques + chart."""
    run = _verify_run_owner(run_id, db, current_user)
    # Migration lazy : recalcule les métriques si les clés Pro sont absentes
    if _run_metrics_stale(run.metrics):
        recompute_run_metrics(run_id, db)
        db.commit()
        db.refresh(run)
    return run


@router.get("/{run_id}/trades", response_model=TradesPaginated)
def get_run_trades(
    run_id: str,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retourne les trades d'un run avec pagination."""
    run = _verify_run_owner(run_id, db, current_user)

    total = db.query(Trade).filter(Trade.run_id == run_id).count()
    trades = (
        db.query(Trade)
        .filter(Trade.run_id == run_id)
        .order_by(Trade.close_time)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return TradesPaginated(
        items=[TradeOut.model_validate(t) for t in trades],
        total=total,
        page=page,
        per_page=per_page,
        pages=math.ceil(total / per_page) if total else 0,
    )


@router.get("/{run_id}", response_model=RunDetail)
def get_run(run_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    run = _verify_run_owner(run_id, db, current_user)

    # Migration lazy : recalcule les métriques si les clés Pro sont absentes
    if _run_metrics_stale(run.metrics):
        recompute_run_metrics(run_id, db)
        db.commit()
        db.refresh(run)

    trades = db.query(Trade).filter(Trade.run_id == run_id).order_by(Trade.close_time).all()

    return RunDetail(
        **RunOut.model_validate(run).model_dump(),
        trades=[TradeOut.model_validate(t) for t in trades],
    )


@router.delete("/{run_id}", status_code=204)
def delete_run(run_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    run = _verify_run_owner(run_id, db, current_user)
    variant_id = run.variant_id
    # Supprimer les trades associés d'abord
    db.query(Trade).filter(Trade.run_id == run_id).delete()
    db.delete(run)
    db.flush()

    # Recalculer les métriques agrégées
    variant = db.query(Variant).filter(Variant.id == variant_id).first()
    recompute_variant_metrics(variant_id, db)
    if variant:
        recompute_strategy_metrics(variant.strategy_id, db)
    db.commit()


def _is_mt5_report(content: bytes) -> bool:
    """Detect whether an Excel file is an MT5 report.

    Scans cell values for MT5-specific markers (section headers like
    'Positions', 'Deals', 'Orders', and known MT5 column patterns).
    """
    import io as _io
    from openpyxl import load_workbook

    try:
        wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        # Only scan the first 20 rows to keep it fast
        markers = {"positions", "deals", "orders"}
        mt5_cols = {"symbole", "symbol", "volume", "profit", "commission", "echange", "swap"}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 20:
                break
            cells = {str(c).strip().lower() for c in row if c is not None}
            # A marker row ("Positions", "Deals", etc.)
            if cells & markers:
                wb.close()
                return True
            # A header row with MT5-typical columns
            if len(cells & mt5_cols) >= 3:
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


def _read_excel_smart(buf) -> "pd.DataFrame":
    """Read an Excel file, auto-detecting the real data table.

    Many trading platforms export Excel files with metadata/summary rows
    before the actual data table. This function:
    1. Tries a naive pd.read_excel first — if the first row looks like
       real columnar data (no "Unnamed" columns dominating), returns it.
    2. Otherwise, scans the raw rows to find the most likely header row
       (the row with the most non-empty cells) and builds a DataFrame
       from there.
    """
    import pandas as pd
    from copy import deepcopy
    from io import BytesIO

    raw = buf.read()

    # --- Attempt 1: naive read ---
    try:
        df_naive = pd.read_excel(BytesIO(raw), nrows=5)
        cols = list(df_naive.columns.astype(str))
        unnamed = sum(1 for c in cols if c.startswith("Unnamed"))
        # If most columns have real names, this is a clean file — return as-is
        if len(cols) >= 2 and unnamed <= len(cols) // 2:
            return df_naive
    except Exception:
        pass

    # --- Attempt 2: scan for the real header row ---
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Fichier Excel vide")

    # Score each row: row with the most non-empty, string-like cells wins
    best_idx, best_score = 0, 0
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        # Prefer rows where most cells look like text headers (not numbers)
        text_cells = sum(1 for c in non_empty if isinstance(c, str))
        score = len(non_empty) + text_cells
        if score > best_score:
            best_score = score
            best_idx = i

    header_row = rows[best_idx]
    # Strip None trailing columns
    while header_row and header_row[-1] is None:
        header_row = header_row[:-1]

    ncols = len(header_row)
    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(header_row)]

    # Collect up to 5 data rows after header
    data_rows = []
    for row in rows[best_idx + 1:]:
        trimmed = row[:ncols]
        non_empty = [c for c in trimmed if c is not None and str(c).strip()]
        if len(non_empty) < 2:
            break
        data_rows.append(list(trimmed))
        if len(data_rows) >= 5:
            break

    if not data_rows:
        return pd.DataFrame([list(header_row[:ncols])], columns=headers)

    df = pd.DataFrame(data_rows, columns=headers)
    return df


@router.post("/auto-mapping")
async def auto_mapping(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Utilise un LLM pour mapper automatiquement les colonnes d'un fichier (CSV, Excel, XML)."""
    import pandas as pd
    import io as _io
    from services.llm_mapper import auto_map_columns

    content = await file.read()
    filename = (file.filename or "").lower()

    # Pour les fichiers Excel, vérifier d'abord si c'est un rapport MT5
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        if _is_mt5_report(content):
            return {"detected_format": "mt5", "columns": [], "mapping": {}}

    # Détecter le format et lire les premières lignes
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = _read_excel_smart(_io.BytesIO(content))
        elif filename.endswith(".xml"):
            df = pd.read_xml(_io.BytesIO(content))
            df = df.head(5)
        else:
            # Fallback CSV (détecte le séparateur automatiquement)
            sample = content[:4096].decode("utf-8", errors="replace")
            sep = ";" if sample.count(";") > sample.count(",") else ","
            df = pd.read_csv(_io.BytesIO(content), nrows=5, sep=sep)
        df.columns = df.columns.str.strip()
    except Exception:
        ext = filename.rsplit(".", 1)[-1] if "." in filename else "inconnu"
        raise HTTPException(400, f"Impossible de lire le fichier .{ext}. Vérifiez que le format est correct.")

    columns = list(df.columns)
    sample_rows = df.head(3).fillna("").to_dict(orient="records")
    # Convertir les valeurs non-sérialisables en string
    for row in sample_rows:
        for k, v in row.items():
            if not isinstance(v, (str, int, float, bool, type(None))):
                row[k] = str(v)

    try:
        mapping = await auto_map_columns(columns, sample_rows)
    except RuntimeError:
        raise HTTPException(503, "Service IA indisponible. Vérifiez la configuration du serveur.")
    except Exception:
        raise HTTPException(502, "Erreur lors de l'analyse automatique. Réessayez ou utilisez le mode manuel.")

    return {"mapping": mapping, "columns": columns}


@router.post("/preview")
async def preview_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Preview parsing of an MT5 Excel file: column mapping + first trades."""
    content = await file.read()
    filename = (file.filename or "").lower()
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")

    if not is_excel:
        raise HTTPException(400, "Preview disponible uniquement pour les fichiers Excel MT5")

    result = preview_mt5_excel(content)
    return result


@router.post("/import", response_model=RunImportResponse)
async def import_csv(
    variant_id: str = Form(...),
    label: str = Form(...),
    type: str = Form(...),
    initial_balance: float | None = Form(None),
    currency: str | None = Form(None),
    timeframe: str | None = Form(None),
    file: UploadFile = File(...),
    column_mapping: str | None = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import CSV de trades pour créer un Run.

    Étapes :
    1. Parser le CSV
    2. Appliquer le mapping de colonnes
    3. Valider les trades
    4. Détecter les overlaps de dates (warning sans blocage)
    5. Calculer les métriques
    6. Persister Run + Trades + Metrics
    """
    # Vérifier que la variante existe et appartient à l'utilisateur
    variant = db.query(Variant).filter(Variant.id == variant_id).first()
    if not variant:
        raise HTTPException(404, "Variante introuvable")
    strategy = db.query(Strategy).filter(Strategy.id == variant.strategy_id, Strategy.user_id == current_user.id).first()
    if not strategy:
        raise HTTPException(404, "Variante introuvable")

    # Valider le type de run
    if type not in ("backtest", "forward", "live"):
        raise HTTPException(400, "Type doit être backtest, forward ou live")

    # Parser le mapping si fourni
    mapping = None
    if column_mapping:
        try:
            mapping = json.loads(column_mapping)
        except json.JSONDecodeError:
            raise HTTPException(400, "column_mapping n'est pas un JSON valide")

    # Étape 1-3 : Parser et valider le fichier
    content = await file.read()
    filename = (file.filename or "").lower()
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")

    if is_excel:
        trades_data, parse_errors, detected_balance, detected_currency = parse_mt5_excel(content)
    else:
        trades_data, parse_errors, detected_balance, detected_currency = parse_csv(content, mapping)

    # Résoudre la balance initiale : formulaire > CSV > défaut
    resolved_balance = initial_balance or detected_balance or 10000.0
    # Résoudre la currency : formulaire > CSV > défaut
    resolved_currency = currency or detected_currency or "USD"
    currency_source = "detected" if detected_currency else "default"
    if currency:
        currency_source = "manual"

    warnings: list[str] = []
    if parse_errors:
        warnings.extend(parse_errors)

    if not trades_data:
        raise HTTPException(400, f"Aucun trade valide trouvé. Erreurs : {parse_errors}")

    # Étape 4 : Détecter les overlaps de dates avec les runs existants
    new_start = min(t["open_time"] for t in trades_data)
    new_end = max(t["close_time"] for t in trades_data)

    existing_runs = db.query(Run).filter(Run.variant_id == variant_id).all()
    for existing in existing_runs:
        if existing.start_date and existing.end_date:
            # Vérifier l'overlap : les périodes se chevauchent si start_a <= end_b ET start_b <= end_a
            if new_start.date() <= existing.end_date and existing.start_date <= new_end.date():
                warnings.append(
                    f"Overlap de dates détecté avec le run '{existing.label}' "
                    f"({existing.start_date} → {existing.end_date})"
                )

    # Étape 5 : Calculer les métriques
    metrics = compute_metrics(trades_data, initial_balance=resolved_balance)

    # Auto-detect pairs from trades
    detected_pairs = sorted({t["symbol"] for t in trades_data if t.get("symbol")})

    # Étape 6 : Persister
    # Dates calculées automatiquement depuis les trades
    run = Run(
        variant_id=variant_id,
        label=label,
        type=type,
        start_date=new_start.date(),
        end_date=new_end.date(),
        initial_balance=resolved_balance,
        currency=resolved_currency,
        currency_source=currency_source,
        pairs=detected_pairs or None,
        timeframe=timeframe or None,
        metrics=metrics,
    )
    db.add(run)
    db.flush()  # Pour obtenir run.id

    for t in trades_data:
        trade = Trade(run_id=run.id, **t)
        db.add(trade)

    db.flush()

    # Recalculer les métriques agrégées variant + stratégie
    recompute_variant_metrics(variant_id, db)
    recompute_strategy_metrics(variant.strategy_id, db)

    db.commit()
    db.refresh(run)

    return RunImportResponse(
        run_id=run.id,
        nb_trades_imported=len(trades_data),
        warnings=warnings,
        metrics=metrics,
        initial_balance=resolved_balance,
        currency=resolved_currency,
    )
